import xml.etree.ElementTree as ET
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import load_workbook
from datetime import datetime
from unidecode import unidecode
import math
import tkinter as tk
from tkinter import filedialog, simpledialog

# source_directory = r"N:\NEFRODOR\Irlik\Tutaj zapisywać pobrane badania w formacie xml"
# destination_directory = r"N:\NEFRODOR\Irlik\Wyniki testowe"
destination_folder_path = r"N:\NEFRODOR\Pacjenci dializowani"


def get_custom_weight(saved_weight):
    if saved_weight:
        message = f"Naciśnij Enter, aby użyć zapisanej masy ciała {saved_weight}kg, lub wprowadź nową wartość masy ciała: "
    else:
        message = "Naciśnij Enter, aby użyć domyślnej masy ciała 70kg, lub wprowadź nową wartość masy ciała: "

    return simpledialog.askfloat("Custom Weight", message)


def try_float(value):
    try:
        return float(value)
    except ValueError:
        return value


def xml_to_matrix(xml_content):
    root = ET.fromstring(xml_content)
    matrix = []

    for row in root.iter('ROW'):
        matrix_row = []
        for i in range(1, 17):  # Start from 1 instead of 0 to skip the first column
            column_name = f'Column{i}'
            cell = row.find(column_name)
            if cell is not None and cell.text:
                matrix_row.append(cell.text)
            else:
                matrix_row.append('')

        matrix.append(matrix_row)

    # Transpose the matrix
    transposed_matrix = list(map(list, zip(*matrix)))

    # Filter out empty columns
    non_empty_columns = [column for column in transposed_matrix if any(cell != '' for cell in column)]

    # Transpose the matrix back
    filtered_matrix = list(map(list, zip(*non_empty_columns)))

    # Filter out rows where all elements are equal
    final_matrix = [row for row in filtered_matrix if not all(cell == row[0] for cell in row)]

    return final_matrix


def extract_results_woemp_from_xls(matrix, custom_weight):
    results = []
    for i in tests:
        test_found = False
        for row in matrix:
            if row[0] == i:
                results.append(row[3:10][::-1])  # Reverse the list before appending
                test_found = True
        if not test_found:
            results.append(["-"])

    results = [[try_float(cell) for cell in row] for row in results]

    results_woemp = [[elem for elem in inner_list if elem != ''] if len(inner_list) > 0 else [None] for inner_list
                     in results if inner_list]
    print(results_woemp)
    serum_iron_row = tests.index('Żelazo')
    tibc_row = tests.index('TIBC')
    insert_index_tsat = tests.index('TSAT')
    # insert_index_uibc = tests.index('UIBC')
    if serum_iron_row != -1 and tibc_row != -1:
        for idx in range(len(results_woemp[0])):
            try:
                serum_iron = results_woemp[serum_iron_row][0]
                tibc = results_woemp[tibc_row][0]

                # Calculate TSAT
                tsat = round((serum_iron / tibc) * 100, 1)
                results_woemp[insert_index_tsat][0] = tsat

                # # Calculate UIBC
                # uibc = round(tibc - serum_iron)
                # results_woemp[insert_index_uibc][0] = uibc

            except (TypeError, ZeroDivisionError, IndexError):
                results_woemp[insert_index_tsat][0] = "-"
                # results_woemp[insert_index_uibc][0] = "-"

    urea_row = tests.index('Mocznik w surowicy')
    insert_index_kt_v = tests_P.index('Kt/V (orient.)') - 1

    try:

        post_urea = float(results_woemp[urea_row][1])
        pre_urea = float(results_woemp[urea_row][0])
        weight = custom_weight

        # Define UF and t
        uf = 2
        t = 4

        # Calculate Kt/V
        R = post_urea / pre_urea
        kt_v = -math.log(R - 0.008 * t) + (4 - 3.5 * R) * (uf / weight)
        kt_v = round(kt_v, 2)
        results_woemp[insert_index_kt_v][0] = kt_v
    except (ValueError, ZeroDivisionError, IndexError):
        # Handle cases where any of the necessary values are missing or invalid
        results_woemp[insert_index_kt_v][0] = "-"

    max_length = max(len(sublist) for sublist in results_woemp)  # find the length of the longest sublist
    for i in range(len(results_woemp)):
        if len(results_woemp[i]) < max_length:  # add "-" to fill in missing values
            results_woemp[i] += ["-"] * (max_length - len(results_woemp[i]))

    # Ensure that there are at least two columns in each row of results_woemp
    for row in results_woemp:
        if len(row) < 2:
            row.append("-")

    # Insert results_woemp[0][1] as a separate list with index results_woemp[1]
    results_woemp.insert(1, [results_woemp[0][1]])

    # Insert results_woemp[3][1] as a separate list with index results_woemp[3]
    results_woemp.insert(5, [results_woemp[4][1]])

    # Keep the first element in each row
    results_woemp = [[row[0]] for row in results_woemp]

    return results_woemp


def extract_reference_limits(normal_range_str):
    if '-' in normal_range_str:
        lower, upper = normal_range_str.split('-')
        return float(lower), float(upper)
    elif '>' in normal_range_str:
        lower = normal_range_str.split('>')[1]
        return float(lower), float('inf')
    elif '<' in normal_range_str:
        upper = normal_range_str.split('<')[1]
        return float('-inf'), float(upper)
    elif normal_range_str:
        return normal_range_str, normal_range_str
    else:
        return None, None


def compare_to_reference_range(value, normal_range_str):
    if not normal_range_str or normal_range_str == '':
        return None
    if value == '-':
        return 'missing'
    lower_limit, upper_limit = extract_reference_limits(normal_range_str)

    if isinstance(value, str):
        try:
            value = float(value)
        except ValueError:
            pass

    if isinstance(lower_limit, str) or isinstance(upper_limit, str):
        return 'within' if value == normal_range_str else 'higher'
    elif value < lower_limit:
        if value < 0.75 * lower_limit:
            return 'much lower'
        return 'lower'
    elif value > upper_limit:
        if value > 1.5 * upper_limit:
            return 'much higher'
        return 'higher'


parameters = [['Moczn. przed HD', 'mmol/l', '2.76-8.07'], ['Moczn. po HD', 'mmol/l', '2.76-8.07'],
              ['Kt/V (orient.)', '', '>1.2'], ['Kt/V (dokł.)', '', '>1.2'],
              ['Potas przed HD', 'mmol/l', '3.5-5.1'], ['Potas po HD', 'mmol/l', '3.5-5.1'],
              ['Sód', 'mmol/l', '136-145'], ['Żelazo', 'μmol/l', '5.83-34.5'], ['TIBC', 'μmol/l', '55-75'],
              ['TSAT', '%', '30-50'], ['Ferrytyna', 'μg/l', '30-400'], ['ESA', 'j./tydzień', ''],
              ['Venofer', 'amp./mies.', ''],
              ['Fosfor', 'mmol/l', '0.81-1.45'], ['Wapń calk', 'mmol/l', '2.2-2.55'],
              ['Parathormon', 'pg/ml', '15-65'], ['ALP', 'IU/L', '30-90'],
              ['Calperos', 'g/dobę', ''], ['HGB', 'g/dl', '13.5-17'], ['MCV', 'fL', '78-100'],
              ['WBC', '10*3/μL', '4-10.5'], ['PLT', '10*3/μL', '150-450'],
              ['Kreatynina', 'μmol/l', '44-80'], ['eGFR', 'ml/min/1.73m^2', ''],
              ['Białko całk', 'g/l', '64-83'], ['Albuminy', 'g/l', '35.6-52'],
              ['pH', '', '7.35-7.45'], ['HCO3-', 'mmol/l', '21-27'],
              ['ALT', 'IU/L', '0-41'], ['AST', 'IU/L', '0-40'], ['Anty-HBs', 'IU/L', ''], ['HBsAg', '', 'ujemny'],
              ['Anty-HCV Ab', 'IU/L', 'niereaktywny'], ['Anty-HIV Ab', 'IU/L', 'ujemny'], ['INR', '', '0.9-1.3'],
              ['PT', 's', '11-16'], ['APTT', 's', '26-40']]


tests = ['Mocznik w surowicy', 'Kt/V (orientacyjnie)', 'Kt/V (dokładne)', 'Potas', 'Sód', 'Żelazo', 'TIBC', 'TSAT',
         'Ferrytyna', 'ESA', 'Venofer', 'Fosfor w surowicy',
         'Wapń calkowity w surowicy', 'Parathormon 1-84', 'ALP2L Fosfataza alkaliczna', 'Kalperos', 'HGB', 'MCV', 'WBC',
         'PLT', 'Kreatynina w surowicy', 'eGFR', 'Białko całkowite', 'Albuminy w surowicy', 'pH', 'HCO3act', 'ALT',
         'AST', 'Anty HBs', 'HBsAg', 'Anty HCV', 'Anty HIV combi', 'INR', 'Czas protrombinowy', 'Czas kaolinowo-kefalinowy']

tests_P = [row[0] for row in parameters]
units = [row[1] for row in parameters]
normal = [row[2] for row in parameters]

# List of tests for which manual completion is required
manual_completion_tests = ['Kt/V (dokł.)', 'ESA', 'Venofer', 'Calperos']

# List comprehension to find the indexes of the specified tests in the `parameters` list
manual_completion_indexes = [i for i, row in enumerate(parameters) if row[0] in manual_completion_tests]


# Loop through files in the source_directory
# Check if the file has a .xml extension
def process_xml_file(file_path, file_name):
    # File is an XML file, process it
    # Read the XML file with the specified encoding
    try:
        with open(file_path, 'r', encoding='iso-8859-2') as file:
            xml_content = file.read()
        # Convert the XML content to a matrix
        matrix = xml_to_matrix(xml_content)
        date = matrix[0][3]

        # Convert the date format to DD.MM.YYYY
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        formatted_date = date_obj.strftime('%d.%m.%Y')

        # Your list of column headings
        column_headings = ['Parametr', 'Jedn.', 'Norma', 'Wyniki']

        # Your list of columns, each as a list
        columns = [tests_P, units]
        normal_range = [normal]

        # Define the fill patterns for headers and alternate rows
        column_header_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
        row_header_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        normal_column_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        alternate_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        alternate_column_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        row_header_fill_from_hct = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
        alternate_column_fill_from_hct = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Define fonts
        units_font = Font(size=7, color="7D7D7D")
        blue_font = Font(color="0000FF")
        dark_blue_font = Font(color="000080")
        red_font = Font(color="FF0000")
        dark_red_font = Font(color="800000")

        # Define the alignment for centering headers and matrix values
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")  # New left alignment

        # Define the borders for separating cells
        thin_border_side = Side(style="thin", color="000000")
        column_border = Border(left=thin_border_side)
        left_border = Border(left=thin_border_side)
        right_border = Border(right=thin_border_side)
        bottom_border = Border(bottom=thin_border_side)
        left_right_bottom_border = Border(left=thin_border_side, right=thin_border_side, bottom=thin_border_side)
        right_bottom_border = Border(right=thin_border_side, bottom=thin_border_side)
        all_sides_border = Border(top=thin_border_side, right=thin_border_side, bottom=thin_border_side,
                                  left=thin_border_side)

        file_found = False

        # Loop through the folders in the destination folder path
        for folder in os.listdir(destination_folder_path):
            # Extract the last name from the folder name
            folder_last_name = unidecode(folder.split()[0].lower())

            # alternatywna opcja z uwzględnieniem imienia lub nie
            # folder_last_name = unidecode(folder.lower())
            # if xls_title.lower() in folder_last_name:

            # Get the title of the xls file
            xls_title = unidecode(os.path.splitext(file_name)[0].lower())

            # Check if the folder name matches with the xls title
            if folder_last_name in xls_title.lower():
                file_found = True
                # Define the xlsx file name and path
                xlsx_file_name = "Wyniki.xlsx"
                xlsx_file_path = os.path.join(destination_folder_path, folder, xlsx_file_name)

                # Check if the xlsx file exists
                if os.path.exists(xlsx_file_path):
                    # Load the existing workbook
                    existing_wb = load_workbook(xlsx_file_path)
                    existing_ws = existing_wb.active

                    # Try to read the saved weight from the file
                    saved_weight = existing_ws['C43'].value

                    # Check if there is a saved weight and offer choices
                    custom_weight = get_custom_weight(saved_weight)

                    # Use the saved/default weight if Enter is pressed or a new weight is not provided
                    if not custom_weight:
                        custom_weight = saved_weight if saved_weight else 70
                    else:
                        # Try to convert the input to a float, if it fails, use the saved/default weight
                        try:
                            custom_weight = float(custom_weight)
                        except ValueError:
                            custom_weight = saved_weight if saved_weight else 70

                    print(f"Użyto masy ciała: {custom_weight}kg")

                    results_woemp = extract_results_woemp_from_xls(matrix, custom_weight)

                    # Save the weight back to the file in cell C43
                    existing_ws['C39'] = custom_weight

                    # Find the last column with "Wyniki"
                    last_col_with_wyniki = None
                    header_row = 5  # Assuming the header is in the 5th row

                    for col in range(1, existing_ws.max_column + 1):
                        cell = existing_ws.cell(row=header_row, column=col)
                        if cell.value == "Wyniki":
                            last_col_with_wyniki = col

                    if last_col_with_wyniki is None:
                        raise ValueError("Could not find the 'Wyniki' column in the header")

                    same_date = False

                    # Iterate through the cells in the fourth row
                    for col in range(1, existing_ws.max_column + 1):
                        cell = existing_ws.cell(row=4, column=col)
                        if cell.value == formatted_date:
                            last_col_with_wyniki = col - 1
                            print(f"Zaktualizowano badania {folder.split('.')[0]} wykonane dnia {formatted_date}")
                            same_date = True

                    # Update the fifth row with the new heading 'Wyniki'
                    updated_heading = 'Wyniki'

                    # Add the 'Wyniki' heading in the fifth row, after the last column
                    cell = existing_ws.cell(row=5, column=last_col_with_wyniki + 1)
                    cell.value = updated_heading
                    cell.font = Font(bold=True)
                    cell.alignment = center_alignment
                    cell.fill = column_header_fill
                    cell.border = all_sides_border

                    # Set the column width for the new 'Wyniki' column
                    column_letter = existing_ws.cell(row=5, column=last_col_with_wyniki + 1).column_letter
                    existing_ws.column_dimensions[column_letter].width = 11

                    last_row = len(columns[0]) + 1

                    # Write the results beneath the newly created header
                    for i, row in enumerate(results_woemp):
                        for j, value in enumerate(row):
                            comparison_result = compare_to_reference_range(value, normal[i])
                            arrow = ""
                            font_color = "000000"
                            is_bold = False

                            if comparison_result == 'lower':
                                arrow = "↓"
                                font_color = "0000FF"
                                is_bold = True
                            elif comparison_result == 'much lower':
                                arrow = "↓↓"
                                font_color = "000080"
                                is_bold = True
                            elif comparison_result == 'higher':
                                arrow = "↑"
                                font_color = "FF0000"
                                is_bold = True
                            elif comparison_result == 'much higher':
                                arrow = "↑↑"
                                font_color = "800000"
                                is_bold = True

                            if i not in manual_completion_indexes:
                                cell = existing_ws.cell(row=i + 6, column=last_col_with_wyniki + 1 + j)
                                if arrow:
                                    cell.value = f"{value}{arrow}"
                                else:
                                    cell.value = value

                            # Apply center alignment to all cells in the column
                            cell = existing_ws.cell(row=i + 6, column=last_col_with_wyniki + 1 + j)
                            cell.alignment = center_alignment
                            cell.border = right_border
                            if i == last_row - 2:
                                cell.border = right_bottom_border

                            # Apply alternate row fill to every other row
                            if i % 2 == 0:
                                cell.fill = alternate_row_fill
                            # Apply white fill for the remaining rows
                            else:
                                cell.fill = white_fill

                            cell.font = Font(color=font_color, bold=is_bold)

                    # Write the date variable in the first row above the new "Wyniki" column
                    existing_ws.cell(row=4, column=last_col_with_wyniki + 1, value=formatted_date)
                    existing_ws.cell(row=4, column=last_col_with_wyniki + 1).alignment = center_alignment
                    existing_ws.cell(row=4, column=last_col_with_wyniki + 1).border = all_sides_border

                    # Save the updated workbook
                    existing_wb.save(xlsx_file_path)
                    if not same_date:
                        print(f"Dodano nowe badania {folder.split('.')[0]} wykonane dnia {formatted_date}")
                    os.remove(file_path)

                    # Save the workbook to a file (test)
                    # new_file_name = os.path.splitext(file_name)[0] + ".xlsx"
                    # new_file_path = os.path.join(destination_directory, new_file_name)
                    # existing_wb.save(new_file_path)
                else:
                    # Create a new workbook and select the active worksheet
                    wb = Workbook()
                    ws = wb.active

                    # Try to read the saved weight from the file
                    saved_weight = ws['C43'].value

                    # Check if there is a saved weight and offer choices
                    custom_weight = get_custom_weight(saved_weight)

                    # Use the saved/default weight if Enter is pressed or a new weight is not provided
                    if not custom_weight:
                        custom_weight = saved_weight if saved_weight else 70
                    else:
                        # Try to convert the input to a float, if it fails, use the saved/default weight
                        try:
                            custom_weight = float(custom_weight)
                        except ValueError:
                            custom_weight = saved_weight if saved_weight else 70

                    print(f"Użyto masy ciała: {custom_weight}kg")

                    results_woemp = extract_results_woemp_from_xls(matrix, custom_weight)

                    # Save the weight back to the file in cell C43
                    ws['C39'] = custom_weight

                    # Set the column width
                    for i in range(len(column_headings)):
                        column_letter = ws.cell(row=1, column=i + 1).column_letter
                        if i == 0:
                            ws.column_dimensions[column_letter].width = 15
                        elif i == 1:
                            ws.column_dimensions[column_letter].width = 7
                        else:
                            ws.column_dimensions[column_letter].width = 11

                    ws.row_dimensions[1].height = 30
                    ws.row_dimensions[2].height = 25
                    ws.row_dimensions[3].height = 25

                    # Write the column headings to the first row of the worksheet and apply column header fill
                    for i, heading in enumerate(column_headings):
                        cell = ws.cell(row=1, column=i + 1)
                        cell.value = heading
                        cell.fill = column_header_fill
                        cell.alignment = center_alignment
                        cell.border = all_sides_border
                        cell.font = Font(bold=True)

                    last_row = len(columns[0]) + 1

                    # Write the columns to the worksheet and apply row header fill, units font, column border, and alternate column fill
                    for i, column in enumerate(columns):
                        for j, value in enumerate(column):
                            cell = ws.cell(row=j + 2, column=i + 1)
                            cell.value = value

                            if i == 1:
                                cell.font = units_font
                                cell.alignment = center_alignment
                                cell.fill = alternate_row_fill if j % 2 == 0 else PatternFill()
                            else:
                                cell.fill = row_header_fill_from_hct if j % 2 == 0 else alternate_column_fill_from_hct

                            if i == 0:
                                cell.border = Border(left=Side(style="thin", color="000000"),
                                                     right=Side(style="thin", color="000000"))
                            elif i == 1:
                                cell.border = right_border

                            # Apply the bottom border to the last row
                            if j == last_row - 2:
                                if i == 0:
                                    cell.border = left_right_bottom_border
                                elif i == 1:
                                    cell.border = right_bottom_border
                                else:
                                    cell.border = bottom_border

                    # Write the normal column to the worksheet with neutral fill
                    for i, value in enumerate(normal):
                        cell = ws.cell(row=i + 2, column=3)
                        cell.value = value
                        cell.alignment = center_alignment
                        cell.border = right_border
                        if i == last_row - 2:
                            cell.border = right_bottom_border

                        # Apply alternate row fill to every other row
                        if i % 2 == 0:
                            cell.fill = alternate_row_fill
                        else:
                            cell.fill = white_fill

                    # Write the matrix to the worksheet
                    for i, row in enumerate(results_woemp):
                        for j, value in enumerate(row):
                            cell = ws.cell(row=i + 2, column=j + 2 + len(columns))
                            cell.alignment = center_alignment  # Move alignment outside the if statement
                            cell.border = right_border
                            if i == last_row - 2:
                                cell.border = right_bottom_border

                            # If the current index i is in the manual_completion_indexes list and the current column is "Wyniki" (j=0),
                            # skip writing the value to avoid overwriting manual completions
                            if i in manual_completion_indexes and j == 0:
                                continue

                            comparison_result = compare_to_reference_range(value, normal[i])

                            if comparison_result == 'lower':
                                cell.font = Font(color="0000FF", bold=True)
                                cell.value = f"{value} ↓"
                            elif comparison_result == 'much lower':
                                cell.font = Font(color="000080", bold=True)
                                cell.value = f"{value} ↓↓"
                            elif comparison_result == 'higher':
                                cell.font = Font(color="FF0000", bold=True)
                                cell.value = f"{value} ↑"
                            elif comparison_result == 'much higher':
                                cell.font = Font(color="800000", bold=True)
                                cell.value = f"{value} ↑↑"
                            else:
                                cell.value = value

                        # Apply alternate row fill to every other row
                        if i % 2 == 0:
                            for j in range(len(row)):
                                cell = ws.cell(row=i + 2, column=j + 2 + len(columns))
                                cell.fill = alternate_row_fill

                    # Insert three additional rows at the top
                    ws.insert_rows(1, 4)

                    # Write "dane szpitala" in the first column of the first row
                    ws.cell(row=1, column=1).value = "S.P SZPITALA KLINICZNEGO NR 1 \n41-808 Zabrze, ul. 3 " \
                                                     "Maja 13-15\nTel.: (032) 37-04-439 (piel), -441 (lek)"
                    cell = ws.cell(row=1, column=1)
                    cell.font = Font(size=6)
                    cell.alignment = Alignment(wrap_text=True)
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

                    # Write "Pododdział dializoterapii" in the third column of the first row
                    ws.cell(row=1, column=3).value = "Pododdział Dializoterapii"
                    cell = ws.cell(row=1, column=3)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(vertical='center')
                    # Write "Karta badań kontrolnych 2023r." in the second row
                    ws.cell(row=2, column=1).value = "Karta badań kontrolnych 2023r."

                    # Merge cells in the second row
                    num_columns = len(column_headings)
                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_columns)

                    # Set the font properties and alignment for the text in the second row
                    cell = ws.cell(row=2, column=1)
                    cell.font = Font(size=14, bold=True)
                    cell.alignment = center_alignment

                    # Create the row for the patient name in the third row
                    # Merge cells in the third row with the patient's name
                    cell = ws.cell(row=3, column=1)
                    cell.border = all_sides_border
                    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_columns)
                    cell.alignment = center_alignment
                    # Add patient's name to the third row
                    ws.cell(row=3, column=1).value = folder
                    # Set the font properties for the patient's name
                    cell.font = Font(size=14, bold=True)

                    # Set the title rows to be printed in each page
                    # ws.print_title_rows = '1:1'

                    # Write the date variable in the fourth row, and merge cells above "Wyniki" and "Po dializie"
                    ws.cell(row=4, column=4, value=formatted_date)
                    ws.cell(row=4, column=4).alignment = center_alignment
                    ws.cell(row=4, column=4).border = all_sides_border

                    # Save the workbook to a file (test)
                    # new_file_name = os.path.splitext(file_name)[0] + ".xlsx"
                    # new_file_path = os.path.join(destination_directory, new_file_name)
                    # wb.save(new_file_path)

                    # Save the new xlsx file to the matching folder
                    wb.save(xlsx_file_path)
                    os.remove(file_path)
                    print(f"Dodano nowy plik \"Wyniki\" dla {folder.split('.')[0]} z badaniami wykonanymi dnia "
                          f"{formatted_date}")

                    break
    except ET.ParseError as e:
        print(f'Error parsing XML file {file_path}: {e}')
    except UnicodeError as e:
        print(f'Error reading XML file with specified encoding {file_path}: {e}')
    finally:
        # print("Skontaktuj się z twórcą skryptu, nr tel. 513 436 210")
        # input("Przyciśnij Enter aby zamknąć")
        pass


def option1_selected():
    label.config(text="Tu bedzie miejsce do wgrywania plików")
    file_paths = filedialog.askopenfilenames()
    if file_paths:
        print("Selected files:")
        for file_path in file_paths:
            file_name = os.path.basename(file_path)
            process_xml_file(file_path, file_name)


def option2_selected():
    label.config(text="Tu bedzie miejsce do pobierania plików")


root = tk.Tk()
root.title("Dializy")
root.geometry("600x400")

# Sidebar frame
sidebar = tk.Frame(root, width=200, bg='lightgrey')
sidebar.pack(side=tk.LEFT, fill=tk.Y)

# Labels for options in sidebar
option1 = tk.Label(sidebar, text="Wybierz pliki", bg='lightgrey', padx=10, pady=5, cursor='hand2')
option1.pack()
option1.bind("<Button-1>", lambda event: option1_selected())

option2 = tk.Label(sidebar, text="Pobierz pliki", bg='lightgrey', padx=10, pady=5, cursor='hand2')
option2.pack()
option2.bind("<Button-1>", lambda event: option2_selected())

# Main content area
main_content = tk.Frame(root, width=400, height=400, bg='white')
main_content.pack_propagate(False)
main_content.pack(expand=True, fill=tk.BOTH)

# Label to display selected option
label = tk.Label(main_content, text="Wybierz opcję z menu", padx=10, pady=10)
label.pack()

root.mainloop()
